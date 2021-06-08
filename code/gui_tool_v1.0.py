
# coding: utf-8

'''
Data reduction tool module
--------------------------
--------------------------
 This is version 1.
Apart from ploting the hysteresis before and after fitting in order to run the data one by one. 
This tool allows the user to save the specimen in a database and either run them later or modify their properties.
Modifying the parameters is either by clicking on parameter sitting after validating the specimen or directly on the database.
Also, the specimen is added from similar possibilities.
To simplify the use, saving plots is offered either as power point or word templates
'''
import sys
import os
import re
import glob
import webbrowser
from openpyxl import load_workbook
import sqlite3
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.backends.backend_qt5agg import NavigationToolbar2QT as NavigationToolbar
import math as m
from math import sqrt
from PyQt5 import uic
from PyQt5.QtCore import *
from PyQt5.QtWidgets import *
from PyQt5.QtGui import *

# # Prepare the input table
#Code is more often read than written.
class Inputable:
    """
    Inputable class used to represent the input table of all specimen/rawdata
    ...

    Attributes
    ----------
    path : file
        The name or the path to the text file containing data to be reduced. 
    inputfile : file, e.g 'input.xls'
        the path or the name of the input file containing parameters of specimen:
        (e.g "Versuchstabelle_LCF").
    filetoimport : widget -> file
        The input table extracted from the excel input file and the raw data.
    inputable : dataframe
        The input table extracted from the excel input file and the raw data.

    Methods
    -------
    ...
    """

    def __init__(self,path,inputfile):
        """
        Parameters
        ----------
        path : str
            The name or the path to the text file containing data to be reduced. 
        inputfile : str
            the path or the name of the input file containing parameters of specimen:
            (e.g "Versuchstabelle_LCF").
        filetoimport : widget -> file
            The input table extracted from the excel input file and the raw data.
        inputable : dataframe
            The input table extracted from the excel input file and the raw data.
        """
       
        self.path = path
        self.inputfile = inputfile
        self.filetoimport = None
        
        self.aRowdata=None
        self.inputable = None
        self.key = None
        self.n_parts = 0
        #self.kraft_limit = 0.091
        self.direct_access = True

    def set_key(self):

        """Initialize: self.key,
            Get the key name of the specimen from the path/name of the raw data.

        The key name allow the access to the input parameters in the input excel file:
        (e.g "Probenbezeichnung" in the excel file "Versuchstabelle_LCF")

        """
        if self.direct_access == False:
            x = glob.glob(os.path.join(self.path, '*.txt'))
            #x.sort(reverse=True)
            keyname = re.split("[_.]", x[0])[-3] #get the key name
            keyname = re.split("[-.]", keyname)

            prefix = keyname.pop(0)
            name =list(keyname)
            name= ''.join(name)
            self.key = prefix+"-"+name # the same format in the excel "Probenbezeichnung"
        else:#in case the name of the file is the key
            x = glob.glob(os.path.join(self.path, '*.txt'))
            #x.sort(reverse=True)#to make surse it is sorted
            prefix = re.split("[-.]", x[0])[1] #get the key name
            self.key = 'AKT3-'+prefix 
            


    def inputs(self):

        """return: Dataframe [:,13 columns+file_name/sheet_name column],
            Extract the key parameters from the excel input file and restructure them into a dataframe.
        """

        input_data = pd.read_excel(self.inputfile, sheet_name=0,header=None,skiprows=9)
        self.inputable = input_data.iloc[:, [1,0,3,4,8,12,15]] # only the useful data
        
        #cleaning and renaming
        self.inputable=self.inputable.astype(object)
        self.inputable=self.inputable.dropna()
        self.inputable.rename(columns={0:'probennr',1:'Probenbezeichnung',3:'Room_temperature',4:'nDurchmesser_Probe',
                                       8:'nBezugslaenge',12:'nZyklenzeit',15:'nZyklenzahl'},inplace=True)
    
        self.inputable.insert(7, "aExt_ist_0",value = None)
        """eps_SD"""
        self.inputable.insert(8, "eps_SD",value = 'Not yet calculated')
        """Zeit_step """
        self.inputable.insert(9, "Zeit_step",value = 0.05)

        n = len(self.inputable)
        """eps fitting range""" 
        self.inputable.insert(10, 'eps_range',n*[[0,0.1]])

        """number of points in one Strain/dehnnung cycle """ 
        self.inputable.insert(11, 'D_points_in1cycle',n*[[4,2]])

        """number of points in one Stress/Spannung cycle""" 
        self.inputable.insert(12, 'S_points_in1cycle',n*[[40,15]])

        """Zeit_step """
        self.inputable.insert(13, "eps_step",value = 0.0002)
        
        """ output file name/sheet name ..."""
        self.inputable['name'] =  self.inputable['Room_temperature'].apply(
            lambda x :str(x))+ '-' +self.inputable['probennr'].apply(lambda x : re.split("[-.]",x)[-1])

        """set the key"""
        self.inputable.set_index('Probenbezeichnung',inplace=True)
        
    def Load_data(self):
        #if( os.path.isdir(self.path)):
        #if( os.path.isfile(self.path)):
        
        x = glob.glob(os.path.join(self.path, '*.txt'))
        self.n_parts = len(x)
        #if self.n_parts==0:
        #print('This folder does not contain any .txt file')
            
        if self.n_parts>0:
            #x.sort(reverse=True)#L.. then N_L... then N_N_L...
            names = ['Zeit','Kraft','Ext_ist', 'M.Weg' ,'Ext-Soll','Regler Soll','Regler Ist',
                                 'Temp. unten','Temp. oben','?','Zyklen']
            self.aRowdata = pd.read_csv(x[0], names = names, delimiter= '\t')
            self.aRowdata.dropna(inplace=True)
            if self.n_parts>1:
                for i,filename in enumerate(x):
                    #print('The following file was loaded: '+ filename)
                    if i >0:
                        names = ['Zeit','Kraft','Ext_ist', 'M.Weg' ,'Ext-Soll','Regler Soll','Regler Ist',
                                         'Temp. unten','Temp. oben','?','Zyklen']

                        N_out = pd.read_csv(filename,names = names,delimiter= '\t')
                        N_out.dropna(inplace=True)
                        self.aRowdata = self.aRowdata.append(N_out)
                        
            self.aRowdata.reset_index(drop=True,inplace=True)
            self.aRowdata =self.aRowdata.astype('float64')
            
            
    def get_inputline(self):
        """
        Display_inputable return: dataframe=self.inputable dataframe,
            The dataframe is displayed in its dataframe format. 
        """  
        #get the input variable from the excelsheet 0
        if self.key is None: self.set_key()

        #run inputs(self) in case self.inputable is empty
        if(self.inputable is None): Inputable.inputs(self)
        self.line = self.inputable.loc[self.key,:]
        return self.line

    def Export_inputable(self,Masterfile):
        """
        Export_inputable export: dataframe=self.inputable dataframe,
            The dataframe isimported to an xls file in order to see and modify it if desired
        """
        #self.Add_IST()
        if os.path.isfile(Masterfile):
            book = load_workbook(Masterfile)
        else:
            writer = pd.ExcelWriter(Masterfile, engine='xlsxwriter')
            writer.save()
            book = load_workbook(Masterfile)
    
        writer = pd.ExcelWriter(Masterfile, engine='openpyxl', mode='a') 
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        
        inputs = [self.inputable]
        sheet_name = re.split("[.]", Masterfile)[0]
        for output in inputs:
            output.to_excel(writer, sheet_name = sheet_name ,index=True,startrow=0,startcol=0,freeze_panes=(1,1))
            writer.save()
            
        """adjust the width of A AND B"""
        worksheet = writer.sheets[sheet_name]
        for col in worksheet.columns:
            worksheet.column_dimensions[col[0].column].width = 20.0
            
        writer.close()

    def Import_inputable(self):
        """
        Import_inputable import: Materfile to self.inputable,
            The excel file is imported to the dataframe attribute self.inputable
        """
        if self.filetoimport == None:
            #print("No path was given!")
            pass
            
        elif os.path.isfile(self.filetoimport):
            sheet_name = re.split("[.]", self.filetoimport)[0]
            input_data = pd.read_excel(self.filetoimport, sheet_name=sheet_name,skiprows=1)

            """replace nan with none"""
            input_data = input_data.astype(object).replace(np.nan, 'None')

            """set the key"""
            input_data.set_index('Probenbezeichnung',inplace=True)
            self.inputable = input_data
            #print('Input table "'+self.w_filetoimport.value+'" was imported succefully!! \n')
            #print('Input table was imported succefully!! \n',self.inputable.head(1))
        else:
            #print('"'+self.w_filetoimport.value+'"'+"(.xls) input file doesn't exist")
            pass

# # Parameters + 50 % hysteresis + Fitting
class Prepare_parameters(Inputable):
    
    def __init__(self,path,inputfile):
        Inputable.__init__(self,path,inputfile)
        self.path = path
        self.inputfile = inputfile
        self.calc_data = None
        self.mid_hysteresis = None
        #if (automatic extraction):     #calls self.Add_IST() which calls self.input() 
        
        self.nDurchmesser_Probe=None
        self.nBezugslaenge=None
        self.nZyklenzeit =None
        self.nExt_ist_0 = None
        self.Zeit_step = None
        self.eps=None
        self.S_range=None
        self.D_range=None
        self.eps_step=None
        
                
        self.rows = None
        self.cycles = None
        self.test_cycles = None
        self.start_half = None
        
        self.sheetname = None
        self.dataname = None
        self.offset = 0
        
    def init_Parameters(self):
        line = self.line.values
        self.sheetname = line[1] #room temperature
        self.nDurchmesser_Probe=line[2]
        self.nBezugslaenge=line[3]
        self.nZyklenzeit =line[4]
        self.test_cycles = line[5]
        self.nExt_ist_0 = line[6]
        self.eps_SD = line[7]#it is not yet calculated
        self.Zeit_step = line[8]
        self.eps=line[9]#guess the right range for quick processing
        self.S_range=line[11]#number of points in one cycle =40+/- 15
        self.D_range=line[10]#number of points in one cycle = 6+/- 2
        self.eps_step=line[12]#increment step 
        self.dataname = line[13] #data name
        
                
    def init_Rowdata(self,ohne_Zyklus=True,ohne_nan=True):
            
        if ohne_Zyklus and ohne_nan:
            #offset1 = len(self.aRowdata.index)
            self.aRowdata = self.aRowdata.dropna()
            #offset = offset1 - len(self.aRowdata.index)
            #if offset>0:
            #print(str(offset)+' lines contains nan values and thus removed') 
            self.aRowdata = self.aRowdata[self.aRowdata['Zyklen']!=0.0]
            self.offset = self.aRowdata.index[0]/(self.nZyklenzeit/self.Zeit_step)
            self.aRowdata.reset_index(drop=True,inplace=True)
            #if self.offset>0:
            #print(str(round(offset,2))+' cycles removed, no load cycles')  
            self.rows = len(self.aRowdata)
            self.cycles = self.rows/(self.nZyklenzeit/self.Zeit_step)
            #print(' The rowdata contains '+str(self.rows) +' lines'+' or '+str(self.cycles)+' cycles')

        elif ohne_Zyklus and not ohne_nan:
            self.aRowdata = self.aRowdata[self.aRowdata['Zyklen']!=0.0]
            #offset = self.aRowdata.index[0]/(self.nZyklenzeit/self.Zeit_step)
            self.aRowdata = self.aRowdata.reset_index(drop=True)
            #if offset>0:
            #print(str(round(offset,2))+' cycles removed, no load cycles')  
            self.rows = len(self.aRowdata.index)
            self.cycles = self.rows/(self.nZyklenzeit/self.Zeit_step)
            #print(' The rowdata contains '+str(self.rows) +' lines'+' and '+str(self.cycles)+' cycles')

        elif ohne_nan and not ohne_Zyklus:
            #offset1 = len(self.aRowdata.index)
            self.aRowdata = self.aRowdata.dropna()
            #offset = offset1 - len(self.aRowdata.index) 
            #if offset>0:
            #print(str(offset)+' cycles removed, no load cycles') 
          
            self.rows = len(self.aRowdata.index)
            self.cycles = self.rows/(self.nZyklenzeit/self.Zeit_step)
            #print(' The rowdata contains '+str(self.rows) +' lines'+' and '+str(self.cycles)+' cycles')

        elif not(ohne_Zyklus and ohne_nan):
            self.rows = len(self.aRowdata.index)
            self.cycles = self.rows/(self.nZyklenzeit/self.Zeit_step)
            #print(' The rowdata contains '+str(self.rows) +' lines'+' and '+str(self.cycles)+' cycles')
            
        if self.line['nZyklenzahl']=='-':
            self.line.at['nZyklenzahl']= self.cycles
            
    def Add_IST(self):
        """
        Add_IST insert: a float column to the self.inputable dataframe,
            It is the initial extensometer value: EXT_IST_0 from the raw data file.
            It requires the attribute self.inputable to be already filled
        """

        """aExt_ist_0"""
        #print('The data reduction of ',self.key,' is starting!')
        aExt_ist_0 = self.inputable['aExt_ist_0'][self.key]
        if  aExt_ist_0 is None :
            value = self.aRowdata['Ext_ist'].iloc[0]
            self.inputable.loc[self.key,'aExt_ist_0'] = value
            self.nExt_ist_0 = value

    def stress_strain(self):

        #inputs
        aZeit=self.aRowdata['Zeit'] # time step
        aKraftinkN=self.aRowdata['Kraft'] #force in kN
        aExt_ist=self.aRowdata['Ext_ist'] #extensometer in state 0
        #aZyklen=self.aRowdata['Zyklen'] #cycles
    
        #conversion
        aKraftinN=aKraftinkN*1000.0 # conversion to N

        #nQuerschnittsflaeche
        nQuerschnittsflaeche=m.pi*(self.line['nDurchmesser_Probe'])**2/4.0 #cross sectonnal area

        #stress and strain calculation
        aSpannung=aKraftinN/nQuerschnittsflaeche # stress calculation
        aDehnung=(aExt_ist-self.line['aExt_ist_0'])/(self.line['nBezugslaenge']+self.line['aExt_ist_0']) # strain calculation

        #output dataframe
        self.calc_data = pd.concat([aZeit,(aZeit - aZeit[0])/self.line['nZyklenzeit'] ,aSpannung,aDehnung],
                                   axis=1, keys=['Zeit','Cycle','Spannung','Dehnung'])

    def midata_hysteresis(self,plot =True,save = False):
        
        if(self.test_cycles=='-'):
            
            #print('---------------------------------------------Warning------------------------------------------------')
            #print('The number of the test cycles is not available and thus estimated (only 20% of the raw data is considered)')
            #print('----------------------------------------------------------------------------------------------------')
            
            onecycle = int(self.line['nZyklenzeit']/self.line['Zeit_step']) #one cycle number of evaluated points
            start = int(self.rows//10) # the evaluated point to avoid rubish
            end = start + onecycle + 1
            self.start_half = int(start//onecycle)

            #'hysteresis'
            self.mid_hysteresis = self.calc_data.iloc[start:end]

            if plot:
                self.mid_hysteresis.plot(x='Dehnung',y='Spannung',title="The 10% raw data hysterisis")
                plt.xlabel(r'$Dehnung (\epsilon [-])$')
                plt.ylabel(r'$Spannung ( s [MPa])$')
                plt.legend()
                plt.show()
            elif save:
                plt.savefig('The 10% raw data hysterisis.png')
                
        else:
            half = int(self.line['nZyklenzahl']//2)
            self.start_half = half
            
            self.mid_hysteresis = self.calc_data[(self.calc_data['Cycle']>=half) & (self.calc_data['Cycle']<half+1)]
            if plot:
                self.mid_hysteresis.plot(x='Dehnung',y='Spannung',title="50%_Dehnung_Spannung")
                plt.xlabel(r'$Dehnung (\epsilon [-])$')
                plt.ylabel(r'$Spannung ( s [MPa])$')
                plt.legend()
            elif save:
                plt.savefig('the 50% hysterisis according to the testing data.png')
    
    """ The Ramer-Douglas-Peucker algorithm roughly ported from the pseudo-code
    provided by http://en.wikipedia.org/wiki/Ramer-Douglas-Peucker_algorithm"""
    def distance(self,a, b):
        return  sqrt((a[0]  - b[0]) ** 2 + (a[1] - b[1]) ** 2)
    #
    def point_line_distance(self,point, start, end):
        if (start[0] == end[0]) and (start[1] == end[1]):
            return self.distance(point, start)
        else:
            n = abs( (end[0] - start[0]) * (start[1] - point[1]) - (start[0] - point[0]) * (end[1] - start[1]) )
            d = sqrt( (end[0] - start[0]) ** 2 + (end[1] - start[1]) ** 2)
            return n / d
    #
    def rdp(self,points, epsilon):
        """Reduces a series of points to a simplified version that loses detail,
        but maintains the general shape of the series."""
        dmax = 0.0
        index = 0
        for i in range(1, len(points) -1):
            d = self.point_line_distance(points[i], points[0], points[-1])
            if d > dmax:
                index = i
                dmax = d
        if dmax >= epsilon:
            results = self.rdp(points[:index+1], epsilon)[:-1] + self.rdp(points[index:], epsilon)
        else:
            results = [points[0], points[-1]]
        return results
    
    def data_onecyle(self):
        
        if self.mid_hysteresis is None:
            self.midata_hysteresis(plot =False,save = False)
        #'Spannung'
        pts_Spannung = np.array([[i,j] for i,j in zip(self.mid_hysteresis['Cycle'], self.mid_hysteresis['Spannung'])])

        #'Dehnung'
        pts_Dehnung = np.array([[i,j] for i,j in zip(self.mid_hysteresis['Cycle'], self.mid_hysteresis['Dehnung'])])

        return pts_Spannung,pts_Dehnung
    
    def eps_fit(self):

        S_pts,D_pts = self.data_onecyle()
        # for quick fitting stress tend to have a biggger values
        eps_S = np.arange(self.eps[1],self.eps[0]+2*self.eps_step,-self.eps_step)
        eps_result = np.ones(2)

        for eps_S in eps_S:
            S  = np.array(self.rdp(S_pts,eps_S)).size/2#size of reduced points according to chosen epsilon
            if((S<=self.S_range[0]+self.S_range[1])& (S>=self.S_range[0]-self.S_range[1]) ):
                eps_result[0] = eps_S#the smallest value (i.e [0]) epsilon providing the wanted range of points
                break
        # for quick fitting strain tend to have smaller values
        eps_D = np.arange(self.eps[0]+self.eps_step,self.eps[1]+self.eps_step,self.eps_step)
        for eps_D in eps_D:
            D = np.array(self.rdp(D_pts,eps_D)).size/2
            if ((D<=self.D_range[0]+self.D_range[1]) & (D>=self.D_range[0]-self.D_range[1])):
                eps_result[1] = eps_D
                break
        return eps_result
    
    def Unique_cycle(self,rlts_Dehnung,rlts_Spannung):
    
        both_cycles = np.concatenate((rlts_Spannung[:,0],rlts_Dehnung[:,0]))
        unique_cycle = np.unique(both_cycles)

        return unique_cycle
    
    def data_reduction(self,i_cycle):
        
        rows = self.fragments[i_cycle]
        data_chip = self.calc_data.iloc[rows[0]:rows[1]]
        cycle = data_chip['Cycle']

        #'Spannung'
        pts_Spannung = np.array([[i,j] for i,j in zip(cycle, data_chip['Spannung'])])
        rlts_Spannung = np.array(self.rdp(pts_Spannung,self.eps_SD[0]))
        self.reduced_stress = self.reduced_stress.append(pd.DataFrame(rlts_Spannung,columns =['Cycle','Spannung']))

        #'Dehnung'
        pts_Dehnung = np.array([[i,j] for i,j in zip(cycle, data_chip['Dehnung'])])
        rlts_Dehnung = np.array(self.rdp(pts_Dehnung,self.eps_SD[1]))
        self.reduced_strain = self.reduced_strain.append(pd.DataFrame(rlts_Dehnung ,columns =['Cycle','Dehnung']))

        #'Dehnung' to 'Spannung'
        unique_cycle = self.Unique_cycle(rlts_Dehnung,rlts_Spannung)
        reduced_rawdata = data_chip[data_chip['Cycle'].isin(unique_cycle)]#isin retrns the matching indexes
        Spannung_Dehnung = np.array([[i,j,k] for i,j,k in zip(reduced_rawdata['Cycle'],
                                    reduced_rawdata['Spannung'],reduced_rawdata['Dehnung'])])
        self.hysteresis = self.hysteresis.append(pd.DataFrame(Spannung_Dehnung,
                                                      columns =['Cycle','Spannung','Dehnung']))
        #pull out of the loop later for better performac!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
        start = self.fragments[0][0]
        end = self.fragments[-1][1]
        self.not_reduced_data = self.calc_data.iloc[start:end]
        
    
    def run_reduced_hyst(self,plot=True):
        i_cycle = int(self.start_half)
        rows = self.fragments[i_cycle]
        data_chip = self.calc_data.iloc[rows[0]:rows[1]+1]
        cycle = data_chip['Cycle']

        #'Spannung'
        pts_Spannung = np.array([[i,j] for i,j in zip(cycle, data_chip['Spannung'])])
        rlts_Spannung = np.array(self.rdp(pts_Spannung,self.eps_SD[0]))
        reduced_stress = pd.DataFrame(rlts_Spannung,columns =['Cycle','Spannung'])

        #'Dehnung'
        pts_Dehnung = np.array([[i,j] for i,j in zip(cycle, data_chip['Dehnung'])])
        rlts_Dehnung = np.array(self.rdp(pts_Dehnung,self.eps_SD[1]))
        reduced_strain = pd.DataFrame(rlts_Dehnung ,columns =['Cycle','Dehnung'])

        #'Dehnung' to 'Spannung'
        unique_cycle = self.Unique_cycle(rlts_Dehnung,rlts_Spannung)
        reduced_rawdata = data_chip[data_chip['Cycle'].isin(unique_cycle)]#isin retrns the matching indexes
        Spannung_Dehnung = np.array([[i,j,k] for i,j,k in zip(reduced_rawdata['Cycle'],
                                    reduced_rawdata['Spannung'],reduced_rawdata['Dehnung'])])
        
        self.mid_reduced_hysteresis = pd.DataFrame(Spannung_Dehnung,columns =['Cycle','Spannung','Dehnung',])
        
        start = self.fragments[0][0]
        end = self.fragments[-1][1]
        self.not_reduced_data = self.calc_data.iloc[start:end]
    
        if plot:
            self.mid_reduced_hysteresis.plot.scatter(x='Dehnung',y='Spannung',
                                                     c='DarkBlue',title="reduced_50%_Dehnung_Spannung")
            plt.xlabel(r'$Dehnung (\epsilon [-])$')
            plt.ylabel(r'$Spannung ( s [MPa])$')
            plt.show()
    
    """Export"""
    def outputfile(self,Output_file_name):
        if os.path.isfile(Output_file_name):
            book = load_workbook(Output_file_name)
        else:
            writer = pd.ExcelWriter(Output_file_name, engine='xlsxwriter')
            writer.save()
            book = load_workbook(Output_file_name)

        writer = pd.ExcelWriter(Output_file_name, engine='openpyxl', mode='a') 
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        return writer
    
    def Export(self,Output_file_name):
        sheetname = str(self.sheetname)
        writer = self.outputfile(Output_file_name+'\\'+ sheetname+'_reduced.xlsx')
        for output in self.output:
            output.to_excel(writer, sheet_name = sheetname ,index=False,startrow=2,startcol=self.shift)
            worksheet = writer.sheets[sheetname]
            worksheet.cell(row=1, column=self.shift+2).value = self.dataname
            writer.save()
            writer.close()
            self.shift+=4
		#print("the reduced data was exported ")
		#print("------------------------------\n ")
	
# # Data reduction process
class Pipeline(Prepare_parameters):
    
    def __init__(self,path,inputfile):
        
        Prepare_parameters.__init__(self,path,inputfile) #~4 seconds for ~ 10**6

        """Parameters"""
        self.direct_access= True
        self.line = self.get_inputline()#and gives the full parameters of this specimen 
        self.init_Parameters() #~0.1 second
        self.Load_data()
        self.init_Rowdata() #~4 second
        self.Add_IST()
        #add it to line
        self.line.at['aExt_ist_0']=self.nExt_ist_0 # it is in inputable but not in line
        
        self.stress_strain() #~0.5 seconds
                
        self.eps_SD = 'Not yet calculated' #fitting parameter of the rdp algo
        self.fragments = None
        self.Fragment=True #weither it is total data redaction or not
        
        """Data reduction output"""
        self.reduced_strain =  pd.DataFrame(columns =['Cycle','Dehnung'])
        self.reduced_stress =  pd.DataFrame(columns =['Cycle','Spannung'])
        self.hysteresis =  pd.DataFrame(columns =['Cycle','Spannung','Dehnung'])
        self.not_reduced_data = None
        self.output =None
        self.frac_point = None
  
    def check_hyst(self):
        self.midata_hysteresis(True,False)
            
    def fit_epsilon(self): 
            self.eps_SD=self.eps_fit()
            #print("The fitting has been done.")
            #print("------------------Diagrams----------------")
            #print("Here is the not reduced 50%/10% hysteris:")
            #self.check_hyst()
            #print("Here is the reduced 50%/10% hysteris:")
            self.check_reduced_hyst()
            #print("------------------------------------------")
    
    def sub_ranges(self):
        
        pts_in_one_cycle = int(self.nZyklenzeit/self.Zeit_step)
        nb_chips = int(self.rows/pts_in_one_cycle)
        chips = []
        for i in range(nb_chips):
            chips.append([i*pts_in_one_cycle,(i+1)*pts_in_one_cycle])
        i +=1
        if i*pts_in_one_cycle < self.rows-1:
            chips.append([i*pts_in_one_cycle,self.rows])
        return chips
    
    def check_reduced_hyst(self):
        
        self.fragments = self.sub_ranges()
        #print('Here is the reduced hysteresis in the middle of the row data -->'+str(self.start_half)+
              #'th cycle out of '+str(self.cycles))
        
        self.run_reduced_hyst(False)
           
    def run_reduction(self):
        
        if  self.eps_SD == 'Not yet calculated' :
            self.fit_epsilon()
        
        pts_in_one_cycle = int(self.nZyklenzeit/self.Zeit_step)
        
        Fragment=self.Fragment
        
        if Fragment:
            self.fragments = self.sub_ranges()
        else:
            inputs = input('enter the range of cycles to be reduced! e.g 0,'+str(self.cycles)+'(max cycles) --> ')
            inputs = re.split("[,.]", inputs)
            self.fragments = [[int(i)*int(pts_in_one_cycle) for i in inputs]]

        #end = self.fragments[-1][1]//pts_in_one_cycle
        #start = self.fragments[0][0]//pts_in_one_cycle
        #print('Data reduction has started between the '+ str(start)+'th cycle and the '+ str(end)+'th cycle...')
        
        nb_fragments = len(self.fragments)

        for i in range(nb_fragments):
            self.data_reduction(i)
            #print(str(round(i*100/(nb_fragments),1))+"% has been done. \n")
                    
        self.output =[self.hysteresis]#[self.reduced_strain,self.reduced_stress,self.hysteresis]
        
    def fract_zeit(self,eps_tol=0.03):
        Dehnung = self.hysteresis['Dehnung']
        frac_Dehnung = np.max(Dehnung)-eps_tol
        frac_point =Dehnung[Dehnung>frac_Dehnung].iloc[0]
        self.frac_point=self.hysteresis[self.hysteresis['Dehnung']==frac_point]
        print('\n The fracture point is defined as: \n', self.frac_point)

    def Export_output(self,Output_file_name):

        self.Export(Output_file_name)
            

# # User plotting interface
class Plot_Window(QDialog):
    def __init__(self, parent=None,scatter=False,half = True,before=True):
        super(Plot_Window, self).__init__()
        if before:
            self.data = None if parent==None else parent.mid_hysteresis
        else:
            self.data = None if parent==None else parent.mid_reduced_hysteresis
            
        self.key = '' if parent==None else parent.key
        self.scatter = scatter
        self.half = half
        
        self.constructor()
        
        # Just some button connected to `plot` method
        #self.button = QPushButton('Plot')
        #self.button1 = QPushButton('Figsize templates')
        #self.button.clicked.connect(self.gui_plot)
        #self.button1.clicked.connect(self.figsize_template)

        # set the layout
        layout = QVBoxLayout()
        layout.addWidget(self.toolbar)
        layout.addWidget(self.canvas)
        #layout.addWidget(self.button)
        #layout.addWidget(self.button1)
        self.setLayout(layout)
        self.setWindowTitle('Plot')
        
    def constructor(self):
        
        ''' To which purpose?  '''
        box = QMessageBox()
        box.setIcon(QMessageBox.Question)
        box.setWindowTitle('Figure size!')
        box.setText('To which purpose?')
        box.setStandardButtons(QMessageBox.Yes|QMessageBox.No|QMessageBox.Cancel)
        Default = box.button(QMessageBox.Yes)
        Default.setText('Default')
        PPT = box.button(QMessageBox.No)
        PPT.setText('PPT')
        Word = box.button(QMessageBox.Cancel)
        Word.setText('Word')
        box.exec_()

        if box.clickedButton() == Default:
        # YES pressed
            # a figure instance to plot on
            self.figure = plt.figure()
            self.afont = {'fontname':'Arial'}
            self.tfont = self.afont#{'fontname':'Arial'}
            self.canvas = FigureCanvas(self.figure)

            # this is the Navigation widget
            # it takes the Canvas widget and a parent
            self.toolbar = NavigationToolbar(self.canvas, self)

            ax = self.figure.add_subplot(111)

        # discards the old graph
        # ax.hold(False) # deprecated, see above

            # plot data
            if not(self.half):
                title="The 10% test data hysterisis of specimen: " +str(self.key) 
            else:
                title="The 50% crack initiation hysterisis of specimen: " +str(self.key) 
                
            if self.scatter:
                self.data.plot.scatter(x='Dehnung',y='Spannung',c='DarkBlue',title=title,grid=1,ax=ax)
                plt.xlabel(r'$Dehnung (\epsilon [-])$',**csfont)
                plt.ylabel(r'$Spannung ( s [MPa])$')

            else:
                self.data.plot(x='Dehnung',y='Spannung',style='b*-',grid = 1,ax=ax)
                plt.xlabel(r'$Dehnung (\epsilon [-])$',**self.afont)
                plt.ylabel(r'$Spannung ( s [MPa])$',**self.afont)
                plt.title(title,**self.tfont)

                
            # refresh canvas
            self.figure.tight_layout()
            self.canvas.draw()
            plt.close()

        elif box.clickedButton() == PPT:
        # NO pressed
            # a figure instance to plot on
            self.figure = plt.figure(figsize=(6.93,4.48))
            self.afont = {'fontname':'Arial'}
            self.tfont = self.afont#{'fontname':'Arial'}
            self.canvas = FigureCanvas(self.figure)
            
            # create an axis
            ax = self.figure.add_subplot(111)

            # discards the old graph
            # ax.hold(False) # deprecated, see above

            # plot data
            if not(self.half):
                title="The 10% test data hysterisis of specimen: " +str(self.key) 
            else:
                title="The 50% crack initiation hysterisis of specimen: " +str(self.key) 
                
            if self.scatter:
                self.data.plot.scatter(x='Dehnung',y='Spannung',c='DarkBlue',title=title,grid=1,ax=ax)
                plt.xlabel(r'$Dehnung (\epsilon [-])$',**csfont)
                plt.ylabel(r'$Spannung ( s [MPa])$')

            else:
                self.data.plot(x='Dehnung',y='Spannung',style='b*-',grid = 1,ax=ax)
                plt.xlabel(r'$Dehnung (\epsilon [-])$',**self.afont)
                plt.ylabel(r'$Spannung ( s [MPa])$',**self.afont)
                plt.title(title,**self.tfont)

            # this is the Navigation widget
            # it takes the Canvas widget and a parent
            self.figure.tight_layout()
            self.canvas = FigureCanvas(self.figure)
            # refresh canvas
            self.canvas.draw()
            self.toolbar = NavigationToolbar(self.canvas, self)
            #self.toolbar.hide()
            plt.close()
            self.toolbar.save_figure()#(u'Save', u'Save the figure', u'filesave', u'save_figure
            self.toolbar.home()
           
        elif box.clickedButton() == Word:
        # Cancel pressed
            # a figure instance to plot on
            self.figure = plt.figure(figsize=(5.34,2.67))
            self.afont = {'fontname':'Arial', 'fontsize':8}
            self.tfont = {'fontname':'Arial', 'fontsize':9}
            # this is the Canvas Widget that displays the `figure`
            # it takes the `figure` instance as a parameter to __init__
            #self.figure.clear()

            # create an axis
            ax = self.figure.add_subplot(111)

            # discards the old graph
            # ax.hold(False) # deprecated, see above

            # plot data
            if not(self.half):
                title="The 10% test data hysterisis of specimen: " +str(self.key) 
            else:
                title="The 50% crack initiation hysterisis of specimen: " +str(self.key) 
                
            if self.scatter:
                self.data.plot.scatter(x='Dehnung',y='Spannung',c='DarkBlue',title=title,grid=1,ax=ax)
                plt.xlabel(r'$Dehnung (\epsilon [-])$',**csfont)
                plt.ylabel(r'$Spannung ( s [MPa])$')

            else:
                self.data.plot(x='Dehnung',y='Spannung',style='b*-',grid = 1,ax=ax)
                plt.xlabel(r'$Dehnung (\epsilon [-])$',**self.afont)
                plt.ylabel(r'$Spannung ( s [MPa])$',**self.afont)
                plt.title(title,**self.tfont)

            # this is the Navigation widget
            # it takes the Canvas widget and a parent
            self.figure.tight_layout()
            self.canvas = FigureCanvas(self.figure)
            self.canvas.draw()
            self.toolbar = NavigationToolbar(self.canvas, self)
            #self.toolbar.hide()
            plt.close()
            self.toolbar.save_figure()#(u'Save', u'Save the figure', u'filesave', u'save_figure
            self.toolbar.home()

             # refresh canvas
            
        
    def gui_plot(self):

        # instead of ax.hold(False)
        #self.figure.clear()

        # create an axis
        ax = self.figure.add_subplot(111)

        # discards the old graph
        # ax.hold(False) # deprecated, see above

        # plot data
        if not(self.half):
            title="The 10% test data hysterisis of specimen: " +str(self.key) 
        else:
            title="The 50% crack initiation hysterisis of specimen: " +str(self.key) 
            
        if self.scatter:
            self.data.plot.scatter(x='Dehnung',y='Spannung',c='DarkBlue',title=title,grid=1,ax=ax)
            plt.xlabel(r'$Dehnung (\epsilon [-])$',**csfont)
            plt.ylabel(r'$Spannung ( s [MPa])$')

        else:
            self.data.plot(x='Dehnung',y='Spannung',style='b*-',grid = 1,ax=ax)
            plt.xlabel(r'$Dehnung (\epsilon [-])$',**self.afont)
            plt.ylabel(r'$Spannung ( s [MPa])$',**self.afont)
            plt.title(title,**self.tfont)

            
        # refresh canvas
        self.figure.tight_layout()
        self.canvas.draw()
        plt.close()


# # Progress bar class

class Progress_bar(QDialog):
    def __init__(self):
        '''Load the ui'''
        QDialog.__init__(self) # Call the inherited classes __init__ method
        self.call = uic.loadUi('../ui/progressbar_2.ui', self) # Load the .ui file
        self.show()


# # Threads of calculation (fitting and reduction)

class reduc_thread(QThread):
    result=pyqtSignal(int)

    def __init__(self, parent=None, **kwargs):
        super(reduc_thread,self).__init__(parent, **kwargs)
        self.specimen = parent
        
    def run(self):
        
        #if  self.specimen.eps_SD is None :
        #self.specimen.fit_epsilon()
        
        pts_in_one_cycle = int(self.specimen.nZyklenzeit/self.specimen.Zeit_step)
        
        Fragment=self.specimen.Fragment
        
        if Fragment:
            self.specimen.fragments = self.specimen.sub_ranges()
        else:
            inputs = input('enter the range of cycles to be reduced! e.g 0,'+str(self.specimen.cycles)+'(max cycles) --> ')
            inputs = re.split("[,.]", inputs)
            self.specimen.fragments = [[int(i)*int(pts_in_one_cycle) for i in inputs]]

        #end = self.specimen.fragments[-1][1]//pts_in_one_cycle
        #start = self.specimen.fragments[0][0]//pts_in_one_cycle
        
        nb_fragments = len(self.specimen.fragments)

        for i in range(nb_fragments):
            self.specimen.data_reduction(i)
            count = int((i+1)*100/nb_fragments)
            self.result.emit(count)
            #print(str(round(i*100/(nb_fragments),1))+"% has been done. \n")
                    
        self.specimen.output =[self.specimen.hysteresis]#[self.specimen.reduced_strain,self.specimen.reduced_stress,self.specimen.hysteresis]
        #print('done!')
        #print(self.specimen.outputpath)
        #self.specimen.Export_output(str(self.specimen.outputpath))#excel
        self.specimen.hysteresis.to_csv(str(self.specimen.outputpath)+'//'+str(self.specimen.key)+'.csv',columns =['Cycle','Spannung','Dehnung'],index = False)
        
    def stop(self):
        self.specimen.inputfolder = None
        self.specimen.testablefile = None
        self.specimen.outputpath = None
        self.specimen.counter = 0
        self.terminate()
        

class fitting_thread(QThread):
    #result=pyqtSignal(int)

    def __init__(self,parent=None, **kwargs):
        super(fitting_thread,self).__init__(parent, **kwargs)
        self.specimen = parent
        
    def run(self):
        S_pts,D_pts = self.specimen.data_onecyle()
        # for quick fitting stress tend to have a biggger values
        #eps_S = np.arange(self.specimen.eps[1],self.specimen.eps[0]+2*self.specimen.eps_step,-self.specimen.eps_step)
        eps_S = np.arange(self.specimen.eps[0]+1*self.specimen.eps_step,self.specimen.eps[1]+1*self.specimen.eps_step,self.specimen.eps_step)
        eps_result = np.ones(2)
        
        for i, eps_S in enumerate(eps_S):
            #maximum = 10     
            #self.result.emit(maximum)
            S  = np.array(self.specimen.rdp(S_pts,eps_S)).size/2#size of reduced points according to chosen epsilon
            if((S<=self.specimen.S_range[0]+self.specimen.S_range[1])& (S>=self.specimen.S_range[0]-self.specimen.S_range[1]) ):
                eps_result[0] = eps_S#the smallest value (i.e [0]) epsilon providing the wanted range of points
                break
        # for quick fitting strain tend to have smaller values
        eps_D = np.arange(self.specimen.eps[0]+self.specimen.eps_step,self.specimen.eps[1]+self.specimen.eps_step,self.specimen.eps_step)
        for i, eps_D in enumerate(eps_D):
            #maximum = 10     
            #self.result.emit(maximum)
            D = np.array(self.specimen.rdp(D_pts,eps_D)).size/2
            if ((D<=self.specimen.D_range[0]+self.specimen.D_range[1]) & (D>=self.specimen.D_range[0]-self.specimen.D_range[1])):
                eps_result[1] = eps_D
                break
                
        self.specimen.eps_SD = eps_result
        self.specimen.line.at['eps_SD'] = eps_result
        #print(eps_result)
        #maximum = 100
        #self.result.emit(maximum)
        
                
    def stop(self):
        self.terminate()
        

# # Input window
class input_window(QDialog):
    def __init__(self,parent,text=''):
        self.parameter = parent
        self.text = text
        QDialog.__init__(self) # Call the inherited classes __init__ method
        self.insert = uic.loadUi('../ui/input_0.ui', self) # Load the .ui file 
        self.current_parameters()
        self.insert.ok.clicked.connect(lambda: self.setparameters(self.text))
        #self.insert.show()
    
    def current_parameters(self):
        '''data parameters'''
            
        if self.text == 'Initial extensometer':
            self.insert.lineEdit2.setText(str(self.parameter.Initial_extensometer))
            self.insert.setToolTip('Extensomter value corresponding to the initial extention after data cleaning')
   
        elif self.text == 'Reference length':
            self.insert.lineEdit2.setText(str(self.parameter.reference_length))
            self.insert.setToolTip('The initial length of the specimen')
            
            
        elif self.text == 'Diameter sample':
            self.insert.lineEdit2.setText(str(self.parameter.Diameter_sample))
            self.insert.setToolTip('The diameter of the specimen')

        elif self.text == 'Cycle time':
            self.insert.lineEdit2.setText(str(self.parameter.Cycle_time))
            self.insert.setToolTip('The time corresponding to one cycle of the specimen test')
            
        elif self.text == 'Number of cycles':
            self.insert.lineEdit2.setText(str(self.parameter.Number_cycles))
            self.insert.setToolTip('The number of performed test cycles until crack initiation')
            
        elif self.text == 'Data time step':
            self.insert.lineEdit2.setText(str(self.parameter.Data_step))
            self.insert.setToolTip('The test time step')

        '''fitting parameters'''
        if self.text == 'Epsilon':
            
            if (self.parameter.eps_SD != 'Not yet calculated'):
                eps_SD = str(round(self.parameter.eps_SD[0],6)) + '&' + str(self.parameter.eps_SD[1])
            else:
                eps_SD = ' example: 0.0018 & 0.0002'
                
            self.insert.lineEdit2.setText(eps_SD)
            self.insert.setToolTip('It is a list of the fitting parameters repectively for stress and strain: e.g. 0.0018 & 0.0002')
            
        elif self.text == 'Number of strain points in one cycle':
            self.insert.lineEdit2.setText(str(self.parameter.strain_points))
            self.insert.setToolTip('An optimal number of strain points in one cycle, the result is varying +/-2')
            
        elif self.text == 'Number of stress points in one cycle':
            self.insert.lineEdit2.setText(str(self.parameter.stress_points))
            self.insert.setToolTip('An optimal number of stress points in one cycle, the result is varying +/-15')
            
        elif self.text == 'Maximum of fitting epsilon tolerance':
            self.insert.lineEdit2.setText(str(self.parameter.max_eps))
            self.insert.setToolTip('The maximum value of the distance between the dataset and the reduced data')
            
        elif self.text == 'Fitting step':
            self.insert.lineEdit2.setText(str(self.parameter.fitting_step))
            self.insert.setToolTip('The step of the method to find the optimal epsilon for the data reduction, or the smalest accepted distance between the data set and the reduced data')
            
                     
    def setparameters(self,txt):
        '''data parameters'''
        
        if self.text == 'Initial extensometer':
            number = self.insert.lineEdit2.text()
            try:
                self.parameter.Initial_extensometer = float(number)
            except Exception:
                QMessageBox.critical(self, 'Inputs error message', "The inserted value should be a number")
            
        elif self.text == 'Reference length':
            
            number = self.insert.lineEdit2.text()
            try:
                self.parameter.reference_length = float(number)
            except Exception:
                QMessageBox.critical(self, 'Inputs error message', "The inserted value should be a number")
            
        elif self.text == 'Diameter sample':
            number = self.insert.lineEdit2.text()
            try:
                self.parameter.Diameter_sample = float(number)
            except Exception:
                QMessageBox.critical(self, 'Inputs error message', "The inserted value should be a number")
            
        elif self.text == 'Cycle time':
            number = self.insert.lineEdit2.text()
            try:
                self.parameter.Cycle_time = int(number)
            except Exception:
                QMessageBox.critical(self, 'Inputs error message', "The inserted value should be an integer")
            
        elif self.text == 'Number of cycles':
            number = self.insert.lineEdit2.text()
            try:
                self.parameter.Number_cycles = int(number)
            except Exception:
                QMessageBox.critical(self, 'Inputs error message', "The inserted value should be an integer")
            
        elif self.text == 'Data time step':
            number = self.insert.lineEdit2.text()
            try:
                self.parameter.Data_step = float(number)
            except Exception:
                QMessageBox.critical(self, 'Inputs error message', "The inserted value should be a number")

        '''fitting parameters'''
        
        if self.text == 'Epsilon':
            number = self.insert.lineEdit2.text()
            if len(np.fromstring(number, dtype=float, sep='&'))==2:
                self.parameter.eps_SD = number
            else:
                QMessageBox.critical(self, 'Inputs error message', "The inserted value should be like 0.0018 & 0.0002 ")
    
        elif self.text == 'Number of strain points in one cycle':
            number = self.insert.lineEdit2.text()
            try:
                self.parameter.strain_points = float(number)
            except Exception:
                QMessageBox.critical(self, 'Inputs error message', "The inserted value should be a number")
            
        elif self.text == 'Number of stress points in one cycle':
            number = self.insert.lineEdit2.text()
            try:
                self.parameter.stress_points = float(number)
            except Exception:
                QMessageBox.critical(self, 'Inputs error message', "The inserted value should be a number")
            
        elif self.text == 'Maximum of fitting epsilon tolerance':
            number = self.insert.lineEdit2.text()
            try:
                self.parameter.max_eps = float(number)
            except Exception:
                QMessageBox.critical(self, 'Inputs error message', "The inserted value should be a number")
            
        elif self.text == 'Fitting step':
            number = self.insert.lineEdit2.text()
            try:
                self.parameter.fitting_step = float(number)
            except Exception:
                QMessageBox.critical(self, 'Inputs error message', "The inserted value should be a number")

        self.insert.close()
            

# # Browse the database file
class database_import(QDialog):
    def __init__(self):
        
        self.database_path = None
        '''Load the ui'''
        QDialog.__init__(self) # Call the inherited classes __init__ method
        self.call = uic.loadUi('../ui/database_0.ui', self) # Load the .ui file
        
        ''' input: browse'''
        self.call.browse1.clicked.connect(self.open_database)#browse folders
        
        '''Validate and set in input'''
        self.call.validate1.clicked.connect(self.validate)#check path
        #self.show()
        
    def validate(self):
        try:
            self.database_path = self.call.lineEdit1.text()
            if self.database_path == '':
                    self.database_path = '../database/newdatabase.db'

            elif os.path.dirname(self.database_path) =='':
                    self.database_path = '../database/'+ self.database_path

            else:
                pass
            self.conn = sqlite3.connect(self.database_path)
            self.c = self.conn.cursor()
            self.c.execute("CREATE TABLE IF NOT EXISTS specimens(name TEXT PRIMARY KEY, number TEXT,temperature TEXT,infolder TEXT,testable TEXT,output TEXT,strain INTEGER, stress INTEGER , epsilonmax REAL, fitstep REAL,eps INTEGER,extensometer REAL,length REAL, diameter REAL, timecycle REAL,numbercycles REAL,timestep REAL)")
            self.c.close()
            self.conn.close()
        except Exception:
            buttonReply = QMessageBox.warning(self,' Error',' Could not open the datbase!')
        self.close()
    
    def open_database(self):
        filename = QFileDialog.getOpenFileName(self, 'Select file',filter="Database(*.db  *)")
        self.call.lineEdit1.setText(filename[0])

#test
def main():
    app = QApplication(sys.argv) # Create an instance of QApplication
    window = database_import()
    #db.connections.close_all()# Create an instance of our class
    #app.aboutToQuit.connect(app.deleteLater)
    app.exec_() # Start the application  
#main()

# # Insert to database class

class insert_window(QDialog):
    def __init__(self,path=None,parent =None):
        
        '''Load the ui'''
        QDialog.__init__(self) # Call the inherited classes __init__ method
        self.call = uic.loadUi('../ui/insert_1.ui', self) # Load the .ui file # Load the .ui file
        self.path = path
        self.specimen = parent

        if self.specimen is None:
            '''initialisation of parameters'''
            '''specimen name'''
            self.specimen_name = ''
            self.specimen_number =''
            self.specimen_temperature= ''

            '''specimen locations'''
            self.main_folder =''
            self.test_table =''
            self.output = ''

            '''data parameters'''  
            self.Initial_extensometer = '-'
            self.reference_length = '-'
            self.Diameter_sample = '-'
            self.Cycle_time = '-'
            self.Number_cycles= '-'
            self.Data_step = 0.05
            '''fitting parameters'''
            self.eps_SD = 'Not yet calculated'
            self.strain_points= 4
            self.stress_points= 40
            self.max_eps= 0.1
            self.fitting_step = 0.0002
        else:
            pass
        
        '''4th input: browse'''
        self.call.browse1.clicked.connect(self.open_inputfolder_on_click)#browse folders
        
        '''5th input: browse'''
        self.call.browse2.clicked.connect(self.open_tablefile_on_click)#browse folders
        
        '''6th input: browse'''
        self.call.browse3.clicked.connect(self.open_outputfolder_on_click)#browse folders
        
        '''fitting parameters'''
        self.call.comboBox1.activated[str].connect(self.dataparameter)
        
        '''Data parameters'''
        self.call.comboBox2.activated[str].connect(self.fittingparameter)
            
        '''add button'''
        self.call.add.clicked.connect(self.add_parameters)
        
        #self.show()
        
    def dataparameter(self, text):
        window = input_window(self,text)
        window.exec_()
        '''data parameters are updated'''  
          
    def fittingparameter(self, text):
        window = input_window(self,text)
        window.exec_()
        '''fitting parameters are updated'''
        
    def get_parameters(self):
        '''specimen name'''
        self.specimen_name = self.call.lineEdit1.text()
        self.specimen_number = self.call.lineEdit2.text()
        self.specimen_temperature = self.call.lineEdit3.text()

        '''specimen locations'''
        self.main_folder = self.call.lineEdit4.text()
        self.test_table = self.call.lineEdit5.text()
        self.output = self.call.lineEdit6.text()
        
    '''add parameters'''
    def add_parameters(self):
        '''database parameters'''
        self.get_parameters()
        '''specimen name'''
        name = self.call.lineEdit1.text()
        number = self.call.lineEdit2.text()
        temperature = self.call.lineEdit3.text()

        '''specimen locations'''
        infolder = self.call.lineEdit4.text()
        testable = self.call.lineEdit5.text()
        output = self.call.lineEdit6.text()
        
        '''fitting parameters'''
        eps = str(self.eps_SD)
        strain= self.strain_points
        stress= self.stress_points
        epsilonmax= self.max_eps
        fitstep = self.fitting_step

        '''data parameters'''  

        extensometer = self.Initial_extensometer
        length = self.reference_length
        diameter = self.Diameter_sample
        timecycle = self.Cycle_time
        numbercycles = self.Number_cycles
        timestep = self.Data_step

        try:
            self.conn = sqlite3.connect(self.path)
            self.c = self.conn.cursor()
            self.c.execute("INSERT INTO specimens (name, number,temperature,infolder,testable,output,strain,stress,epsilonmax,fitstep,eps,extensometer ,length, diameter,timecycle,numbercycles,            timestep) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",(name,number,temperature,infolder,testable,output ,strain, stress, epsilonmax, fitstep,eps ,extensometer ,length, diameter , timecycle,numbercycles,timestep))
            self.conn.commit()
            self.c.close()
            #self.conn.execute("PRAGMA busy_timeout = 1000")
            self.conn.close()
            buttonReply = QMessageBox.information(self,'Successful','Specimen is added successfully to the database.')
            self.close()
            
        except Exception:
            buttonReply = QMessageBox.information(self,'error','Could not be added to database!.')
        
    '''4th input: brows'''
    def open_inputfolder_on_click(self):
        foldername = QFileDialog.getExistingDirectory(self, 'Select directory')
        self.call.lineEdit4.setText(foldername)
        self.infolder =foldername

        
    '''5th input: browse'''
    def open_tablefile_on_click(self):
        filename = QFileDialog.getOpenFileName(self, 'Select file',filter="Excel(*.xls *.xlsx)")
        self.call.lineEdit5.setText(filename[0])
        self.testable =filename[0]
           
    '''6th input: browse'''
    def open_outputfolder_on_click(self):
        foldername = QFileDialog.getExistingDirectory(self, 'Select directory')
        self.call.lineEdit6.setText(foldername)
        self.output = foldername
        
#test        
def main():
    app = QApplication(sys.argv) # Create an instance of QApplication
    window = insert_window()
    #db.connections.close_all()# Create an instance of our class
    #app.aboutToQuit.connect(app.deleteLater)
    app.exec_() # Start the application     

#main()


# # Delete database row

class delete_window(QDialog):
    def __init__(self,path=None):
        
        '''Load the ui'''
        QDialog.__init__(self) # Call the inherited classes __init__ method
        self.call = uic.loadUi('../ui/delete_search_0.ui', self) # Load the .ui file # Load the .ui file
        
        '''Parameters'''
        self.key = None
        self.path = path
        
        '''validate'''
        self.call.validate1.clicked.connect(self.delete_key)
        
        #self.show()
        
    '''key'''
    def delete_key(self):
        '''specimen name'''
        self.key = self.call.lineEdit1.text()
        try:
            self.conn = sqlite3.connect(self.path)
            self.c = self.conn.cursor()
            self.c.execute("DELETE from specimens WHERE name="+'"'+self.key+'"')
            self.conn.commit()
            self.c.close()
            self.conn.close()
            buttonReply = QMessageBox.information(self,'Successful','Specimen is deleted successfully from database.')
            self.close()
        except Exception:
            buttonReply = QMessageBox.information(self,'error','Could not delete specimen from database!.')

#test        
def main():
    app = QApplication(sys.argv) # Create an instance of QApplication
    window = delete_window('../database/newdatabase.db')
    #db.connections.close_all()# Create an instance of our class
    #app.aboutToQuit.connect(app.deleteLater)
    app.exec_() # Start the application     
#main()


# # Search for a specimen

class search_window(QDialog):
    def __init__(self,path=None):
        
        '''Load the ui'''
        QDialog.__init__(self) # Call the inherited classes __init__ method
        self.call = uic.loadUi('../ui/delete_search_0.ui', self) # Load the .ui file # Load the .ui file
        
        '''Parameters'''
        self.key = None
        self.path = path
        
        '''validate'''
        self.call.validate1.clicked.connect(self.search_key)

        
    '''key'''
    def search_key(self):
        '''specimen name'''
        self.key = self.call.lineEdit1.text()

        try:
            self.conn = sqlite3.connect(self.path)
            self.c = self.conn.cursor()
            result = self.c.execute("SELECT * from specimens WHERE name="+'"'+self.key+'"')
            row = result.fetchone()
            serachresult = "name : "+str(row[0])+'\n'+"number : "+str(row[1])+'\n'+"temperature : "+str(row[2])+'\n'+"folder : "+str(row[3])+'\n'+"test table : "+str(row[4])
            buttonReply = QMessageBox.information(self,'Successful',serachresult)
            self.conn.commit()
            self.c.close()
            self.conn.close()
        except Exception:
            buttonReply = QMessageBox.information(self,'error','Could not find specimen in database!.')


#test        
def main():
    app = QApplication(sys.argv) # Create an instance of QApplication
    window = search_window('../database/newdatabase.db')
    window.show()
    #db.connections.close_all()# Create an instance of our class
    #app.aboutToQuit.connect(app.deleteLater)
    app.exec_() # Start the application     
#main()


# # Modify parameters

class parameters_window(QDialog):
    def __init__(self,parent):
        
        '''Load the ui'''
        QDialog.__init__(self) # Call the inherited classes __init__ method
        self.call = uic.loadUi('../ui/parameters_2.ui', self) # Load the .ui file # Load the .ui file
        
        '''specimen'''
        self.specimen = parent
        '''specimen location'''
        self.folder = self.specimen.inputfolder
        self.testable= self.specimen.testablefile 
        self.output = self.specimen.outputpath

        '''specimen name'''
        self.name = self.specimen.key#line['Probenbezeichnung']
        self.number = self.specimen.line['probennr']
        self.temperature =  self.specimen.line['Room_temperature']

        '''data parameters'''  
        self.Initial_extensometer = self.specimen.line['aExt_ist_0']
        self.reference_length = self.specimen.line['nBezugslaenge']
        self.Diameter_sample = self.specimen.line['nDurchmesser_Probe']
        self.Cycle_time = self.specimen.line['nZyklenzeit']
        self.Number_cycles= self.specimen.line['nZyklenzahl']
        self.Data_step = self.specimen.line['Zeit_step']
        '''fitting parameters'''
        self.eps_SD = self.specimen.line['eps_SD']
        self.strain_points= self.specimen.line['D_points_in1cycle'][0]
        self.stress_points= self.specimen.line['S_points_in1cycle'][0]
        self.max_eps= self.specimen.line['eps_range'][1]
        self.fitting_step = self.specimen.line['eps_step']

        """display: name, num,temperature"""
        self.call.label1.setText('This is specimen : ' +str(self.specimen.line['probennr'])+
                             ' of room: '+str(self.specimen.line['Room_temperature'])+'°C')

        
        """default set parameter"""
        self.call.comboBox1.activated[str].connect(self.dataparameter)
        self.call.comboBox2.activated[str].connect(self.fittingparameter)
        
        """refresh"""
        self.call.refresh.clicked.connect(self.setparameters)
        
        """add"""
        self.call.add.clicked.connect(self.add_todatabase)
        #self.show()
        
    def add_todatabase(self):
        try:
            _import = database_import()#import or create dat base
            _import.exec_()
            path = _import.database_path
            assert path is not None
            '''add parameters'''
            self.add_parameters(path)
        except Exception:
            buttonReply = QMessageBox.warning(self,' Error',' Could not find the database!')
        
    def add_parameters(self,path):
        '''database parameters'''
        '''specimen name'''
        path = str(path)
        name = str(self.name)
        number = str(self.number)
        temperature = str(self.temperature)

        '''specimen locations'''
        infolder = str(self.folder)
        testable = str(self.testable)
        output = str(self.output)
        
        '''fitting parameters'''
        eps = str(self.eps_SD)
        strain= int(self.strain_points)
        stress= int(self.stress_points)
        epsilonmax= float(self.max_eps)
        fitstep = float(self.fitting_step)

        '''data parameters'''  

        extensometer = float(self.Initial_extensometer)
        length = float(self.reference_length)
        diameter = float(self.Diameter_sample)
        timecycle =float(self.Cycle_time)
        numbercycles = float(self.Number_cycles)
        timestep = float(self.Data_step)
        
        #print(name,number,temperature,infolder,testable,output ,strain, stress, epsilonmax, fitstep,skiprow ,extensometer ,length, diameter , timecycle,numbercycles,timestep)
        try:
            self.conn = sqlite3.connect(path)
            self.c = self.conn.cursor()
            self.c.execute("INSERT INTO specimens (name, number,temperature,infolder,testable,output,strain,stress,epsilonmax,fitstep,eps,extensometer ,length, diameter,timecycle,numbercycles,timestep) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",(name,number,temperature,infolder,testable,output ,strain, stress, epsilonmax, fitstep,eps,extensometer ,length, diameter , timecycle,numbercycles,timestep))
            self.conn.commit()
            self.c.close()
            #self.conn.execute("PRAGMA busy_timeout = 1000")
            self.conn.close()
            buttonReply = QMessageBox.information(self,'Successful','Specimen is added successfully to the database.')
            self.close()
            
        except Exception:
            buttonReply = QMessageBox.information(self,'error','Could not be added to database (already existing)!')
        
        
    def setparameters(self):
        '''dataparameters'''
        self.specimen.line.at['aExt_ist_0'] = self.Initial_extensometer
        self.specimen.line.at['nBezugslaenge'] = self.reference_length
        self.specimen.line.at['nDurchmesser_Probe'] = self.Diameter_sample
        self.specimen.line.at['nZyklenzeit'] = self.Cycle_time
        self.specimen.line.at['nZyklenzahl'] = self.Number_cycles
        self.specimen.line.at['Zeit_step'] = self.Data_step
        '''fitting parameters'''
        self.specimen.line.at['eps_SD'] = str(self.eps_SD)
        self.specimen.line.at['D_points_in1cycle'] = [self.strain_points,2]
        self.specimen.line.at['S_points_in1cycle'] = [self.stress_points,15]
        self.specimen.line.at['eps_range'] = [0,self.max_eps]
        self.specimen.line.at['eps_step'] = self.fitting_step
        self.specimen.init_Parameters()
        buttonReply = QMessageBox.information(self,' Parameters information', 
                ' The parameters are updated!')


    def dataparameter(self, text):
        window = input_window(self,text)
        window.exec_()
          
    def fittingparameter(self, text):
        window = input_window(self,text)
        window.exec_()


# # Add directly to database

class add_directly_todatabase(QDialog):
    def __init__(self,parent):
        '''window object'''
        QDialog.__init__(self)
        
        '''specimen'''
        self.specimen = parent
        '''specimen location'''
        self.folder = self.specimen.inputfolder
        self.testable= self.specimen.testablefile 
        self.output = self.specimen.outputpath

        '''specimen name'''
        self.name = self.specimen.key#line['Probenbezeichnung']
        self.number = self.specimen.line['probennr']
        self.temperature =  self.specimen.line['Room_temperature']

        '''data parameters'''  
        self.Initial_extensometer = self.specimen.line['aExt_ist_0']
        self.reference_length = self.specimen.line['nBezugslaenge']
        self.Diameter_sample = self.specimen.line['nDurchmesser_Probe']
        self.Cycle_time = self.specimen.line['nZyklenzeit']
        self.Number_cycles= self.specimen.line['nZyklenzahl']
        self.Data_step = self.specimen.line['Zeit_step']
        '''fitting parameters'''
        self.eps_SD = self.specimen.line['eps_SD']
        self.strain_points= self.specimen.line['D_points_in1cycle'][0]
        self.stress_points= self.specimen.line['S_points_in1cycle'][0]
        self.max_eps= self.specimen.line['eps_range'][1]
        self.fitting_step = self.specimen.line['eps_step']
        
    def add_todatabase(self):
        try:
            _import = database_import()#import or create dat base
            _import.exec_()
            path = _import.database_path
            assert path is not None
            '''add parameters'''
            self.add_parameters(path)
        except Exception:
            buttonReply = QMessageBox.warning(self,' Error',' Could not find the database!')
        
    def add_parameters(self,path):
        '''database parameters'''
        '''specimen name'''
        path = str(path)
        name = str(self.name)
        number = str(self.number)
        temperature = str(self.temperature)

        '''specimen locations'''
        infolder = str(self.folder)
        testable = str(self.testable)
        output = str(self.output)
        
        '''fitting parameters'''
        eps = str(self.eps_SD)
        strain= int(self.strain_points)
        stress= int(self.stress_points)
        epsilonmax= float(self.max_eps)
        fitstep = float(self.fitting_step)

        '''data parameters'''  

        extensometer = float(self.Initial_extensometer)
        length = float(self.reference_length)
        diameter = float(self.Diameter_sample)
        timecycle =float(self.Cycle_time)
        numbercycles = float(self.Number_cycles)
        timestep = float(self.Data_step)
        
        #print(name,number,temperature,infolder,testable,output ,strain, stress, epsilonmax, fitstep,skiprow ,extensometer ,length, diameter , timecycle,numbercycles,timestep)
        try:
            self.conn = sqlite3.connect(path)
            self.c = self.conn.cursor()
            self.c.execute("INSERT INTO specimens (name, number,temperature,infolder,testable,output,strain,stress,epsilonmax,fitstep,eps,extensometer ,length, diameter,timecycle,numbercycles,timestep) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",(name,number,temperature,infolder,testable,output ,strain, stress, epsilonmax, fitstep,eps,extensometer ,length, diameter , timecycle,numbercycles,timestep))
            self.conn.commit()
            self.c.close()
            #self.conn.execute("PRAGMA busy_timeout = 1000")
            self.conn.close()
            buttonReply = QMessageBox.information(self,'Successful','Specimen is added successfully to the database.')
            self.close()
            
        except Exception:
            buttonReply = QMessageBox.information(self,'error','Could not be added to database (already existing)!')


# # Database

class database_window(QDialog):
    def __init__(self,head=None,databasepath=None):
        QDialog.__init__(self) # Call the inherited classes __init__ method
        self.setFixedWidth(876)
        self.setFixedHeight(369)
        self.box = uic.loadUi('../ui/box_1.ui', self) # Load the .ui file 
        
        default  = ("Specimen name","Specimen number","Temperature",
                          \
                        'Specimen folder','Test table file', 'Output folder',
                          \
                          "Strain points in 1 cycle", "Stress points in 1 cycle", "Fitting tolerance ", "Fitting step",
                          \
                          "eps_Stress_Srain","Initial extensometer","Reference length", "Diameter sample", "Cycle time", "Number of cycles","Data time step")
        
        self.header = default if head is None else head
        self.Header(self.header)
        self.database_path = databasepath
        '''create database'''
         
        if self.database_path is not None:
            try:
                if self.database_path == '':
                    self.database_path = '../database/newdatabase.db'

                elif os.path.dirname(self.database_path) =='':
                    self.database_path = '../database/'+ self.database_path

                else:
                    pass
                path = self.database_path

                self.conn = sqlite3.connect(path)
                self.c = self.conn.cursor()
                self.c.execute("CREATE TABLE IF NOT EXISTS specimens(name TEXT PRIMARY KEY, number TEXT,temperature TEXT,infolder TEXT,testable TEXT,output TEXT,                strain INTEGER, stress INTEGER , epsilonmax REAL, fitstep REAL,eps TXT,extensometer REAL,length REAL, diameter REAL, timecycle REAL,numbercycles REAL,timestep REAL)")
                self.c.close()
                self.conn.close()
                buttonReply = QMessageBox.information(self,' Information','Database was created in :'+ self.database_path )
            except Exception:
                buttonReply = QMessageBox.warning(self,' Error','Database is not created!')

            '''refresh button'''
            self.box.Load_specimen.clicked.connect(self.loaddata)
            ''''add  button'''
            self.box.add_specimen.clicked.connect(self.insert)
            ''''search button'''
            self.box.search_specimen.clicked.connect(self.search)
            ''''delete button'''
            self.box.delete_specimen.clicked.connect(self.delete)
            
        else:
            self.box.close()

        #self.box.show()
 
    def Header(self,header):
        
        self.box.tableWidget.setHorizontalHeaderLabels(header)
        
        rgb = [QColor(92, 186, 164),QColor(23, 156, 125),
               QColor(76, 153, 178),QColor(31, 130, 192)]
        
        self.color(header,rgb)
        
    def color(self,header,rgb):
        
        for i,text in enumerate(header):
            if i < 3 :
                item = QTableWidgetItem(text)
                item.setBackground(rgb[0])
                self.box.tableWidget.setHorizontalHeaderItem(i,item)
                
            elif i < 6:
                item = QTableWidgetItem(text)
                item.setBackground(rgb[1])
                self.box.tableWidget.setHorizontalHeaderItem(i,item)
            
            elif i < 11:
                item = QTableWidgetItem(text)
                item.setBackground(rgb[2])
                self.box.tableWidget.setHorizontalHeaderItem(i,item)
            
            else:
                item = QTableWidgetItem(text)
                item.setBackground(rgb[3])
                self.box.tableWidget.setHorizontalHeaderItem(i,item)
                
    def loaddata(self):
        try:
            self.connection = sqlite3.connect(self.database_path)
            query = "SELECT * FROM specimens"
            result = self.connection.execute(query)
            self.tableWidget.setRowCount(0)
            for row_number, row_data in enumerate(result):
                self.box.tableWidget.insertRow(row_number)
                for column_number, data in enumerate(row_data):
                    self.box.tableWidget.setItem(row_number, column_number,QTableWidgetItem(str(data)))
            self.connection.close()
        except Exception:
            buttonReply = QMessageBox.warning(self,' Error','Database is not loaded succefully!')
                

    def insert(self):
        window = insert_window(self.database_path)
        window.exec_()

    def delete(self):
        window = delete_window(self.database_path)
        window.exec_()

    def search(self):
        window = search_window(self.database_path)
        window.exec_()
        
#test
def main():
    app = QApplication(sys.argv) # Create an instance of QApplication
    app.setStyle(QStyleFactory.create('Fusion')) # otherwise colors change won't work on windows style.
    window = database_window() # Create an instance of our class
    #app.aboutToQuit.connect(app.deleteLater)
    sys.exit(app.exec_()) # Start the application

#main()

# # Run all from Data base

class runall_thread(QThread):
    result=pyqtSignal(object)
    
    def __init__(self,dataframe = None):
        QThread.__init__(self) # Call the inherited classes __init__ method
        self.dataframe = dataframe
        '''initialisation of parameters'''
        '''specimen name'''
        self.specimen_name = ''
        self.specimen_number =''
        self.specimen_temperature= ''

        '''specimen locations'''
        self.main_folder =''
        self.test_table =''
        self.output = ''

        '''data parameters'''  
        self.Initial_extensometer = '-'
        self.reference_length = '-'
        self.Diameter_sample = '-'
        self.Cycle_time = '-'
        self.Number_cycles= '-'
        self.Data_step = 0.05
        '''fitting parameters'''
        self.eps_SD = '-'
        self.strain_points= 4
        self.stress_points= 40
        self.max_eps= 0.1
        self.fitting_step = 0.0002 
        
    def run(self):
        head = ('name', 'number','temperature','infolder','testable','output','strain','stress','epsilonmax','fitstep','eps','extensometer' ,'length', 'diameter','timecycle','numbercycles','timestep')
        for index,row in self.dataframe.iterrows():
            
            self.index=index
            self.total_rows= len(self.dataframe.index)
           
            '''specimen name'''
            self.specimen_name = row[head[0]]
            self.specimen_number = row[head[1]]
            self.specimen_temperature= row[head[2]]

            '''specimen locations'''
            self.main_folder = row[head[3]]
            self.test_table = row[head[4]]
            self.output = row[head[5]]
            
            '''specimen initialisation'''
            self.specimen = Pipeline(self.main_folder,self.test_table)

            '''fitting parameters'''
            self.strain_points= 4 if (row[head[6]] == '-') else row[head[6]]
            self.stress_points= 40 if (row[head[7]] == '-') else row[head[7]]
            self.max_eps= 0.1 if (row[head[8]] == '-') else row[head[8]]
            self.fitting_step = 0.0002 if (row[head[9]] == '-') else row[head[9]]
            epsilon_toarray = np.fromstring(row[head[10]], dtype=float, sep='&')
            self.eps_SD = 'Not yet calculated' if (len(epsilon_toarray)!=2) else epsilon_toarray

            '''data parameters'''  
            self.Initial_extensometer = row[head[11]]
            self.reference_length = row[head[12]]
            self.Diameter_sample = row[head[13]]
            self.Cycle_time = row[head[14]]
            self.Number_cycles= int(self.specimen.cycles) if (row[head[15]] == '-') else row[head[15]]
            self.Data_step = 0.05 if (row[head[16]] == '-') else row[head[16]]
            
            '''prepare parameters'''
            self.update_parameters()
            
            '''prepare mid data for fitting'''
            self.specimen.midata_hysteresis(plot =False,save = False)
            
            '''fitting + datareduction'''
            self.data_reduction()
            
    def update_parameters(self):
        '''dataparameters'''
        self.specimen.line.at['aExt_ist_0'] = self.Initial_extensometer
        self.specimen.line.at['nBezugslaenge'] = self.reference_length
        self.specimen.line.at['nDurchmesser_Probe'] = self.Diameter_sample
        self.specimen.line.at['nZyklenzeit'] = self.Cycle_time
        self.specimen.line.at['nZyklenzahl'] = self.Number_cycles
        self.specimen.line.at['Zeit_step'] = self.Data_step
        '''fitting parameters'''
        self.specimen.line.at['eps_SD'] = self.eps_SD
        self.specimen.line.at['D_points_in1cycle'] = [self.strain_points,2]
        self.specimen.line.at['S_points_in1cycle'] = [self.stress_points,15]
        self.specimen.line.at['eps_range'] = [0,self.max_eps]
        self.specimen.line.at['eps_step'] = self.fitting_step
        self.specimen.init_Parameters()
        
    def fit_epsilon(self):
        S_pts,D_pts = self.specimen.data_onecyle()
        # for quick fitting stress tend to have a biggger values
        #eps_S = np.arange(self.specimen.eps[1],self.specimen.eps[0]+2*self.specimen.eps_step,-self.specimen.eps_step)
        eps_S = np.arange(self.specimen.eps[0]+1*self.specimen.eps_step,self.specimen.eps[1]+1*self.specimen.eps_step,self.specimen.eps_step)
        eps_result = np.ones(2)
        
        for i, eps_S in enumerate(eps_S):
            #maximum = 10     
            #self.result.emit(maximum)
            S  = np.array(self.specimen.rdp(S_pts,eps_S)).size/2#size of reduced points according to chosen epsilon
            if((S<=self.specimen.S_range[0]+self.specimen.S_range[1])& (S>=self.specimen.S_range[0]-self.specimen.S_range[1]) ):
                eps_result[0] = eps_S#the smallest value (i.e [0]) epsilon providing the wanted range of points
                break
        # for quick fitting strain tend to have smaller values
        eps_D = np.arange(self.specimen.eps[0]+self.specimen.eps_step,self.specimen.eps[1]+self.specimen.eps_step,self.specimen.eps_step)
        for i, eps_D in enumerate(eps_D):
            #maximum = 10     
            #self.result.emit(maximum)
            D = np.array(self.specimen.rdp(D_pts,eps_D)).size/2
            if ((D<=self.specimen.D_range[0]+self.specimen.D_range[1]) & (D>=self.specimen.D_range[0]-self.specimen.D_range[1])):
                eps_result[1] = eps_D
                break
                
        self.specimen.eps_SD = eps_result
        
    def data_reduction(self):
    
        if  self.specimen.eps_SD == 'Not yet calculated' :
            self.fit_epsilon()
        
        pts_in_one_cycle = int(self.specimen.nZyklenzeit/self.specimen.Zeit_step)
        
        Fragment=self.specimen.Fragment
        
        if Fragment:
            self.specimen.fragments = self.specimen.sub_ranges()
        else:
            inputs = input('enter the range of cycles to be reduced! e.g 0,'+str(self.specimen.cycles)+'(max cycles) --> ')
            inputs = re.split("[,.]", inputs)
            self.specimen.fragments = [[int(i)*int(pts_in_one_cycle) for i in inputs]]

        end = self.specimen.fragments[-1][1]//pts_in_one_cycle
        start = self.specimen.fragments[0][0]//pts_in_one_cycle
        
        nb_fragments = len(self.specimen.fragments)

        for i in range(nb_fragments):
            self.specimen.data_reduction(i)
            count = int((i+1)*100/nb_fragments)
            array_to_send = [count,self.specimen_name,self.index,self.total_rows]
            self.result.emit(array_to_send)
            #print(str(round(i*100/(nb_fragments),1))+"% has been done in epecimen. "+str(self.specimen_number)+'('+str(self.index+1)+'th out of:'+str(self.total_rows)+')\n')
                    
        self.specimen.output =[self.specimen.hysteresis]#[self.specimen.reduced_strain,self.specimen.reduced_stress,self.specimen.hysteresis]
        self.output = '.' if (self.output == '') else self.output
        self.specimen.hysteresis.to_csv(self.output+'//'+str(self.specimen_name)+'.csv',columns =['Cycle','Spannung','Dehnung'],index = False)
        #print('done!')
        
    def stop(self):
        self.specimen.counter = 0
        self.terminate()

# # Validate thread

class validate_thread(QThread):
    
    def __init__(self,parent=None, **kwargs):
        super(validate_thread,self).__init__(parent, **kwargs)
        self.specimen = parent
    
    def run(self):
        try:
            Pipeline.__init__(self.specimen,self.specimen.inputfolder,self.specimen.testablefile) # Call the inherited classes __init__ method
        except Exception:
            pass
            #self.stop()
            #buttonReply = QMessageBox.information(self,'Error message', "Paths not valid!")

    def stop(self):
        self.terminate()
        

# # User Interface Class

class datareduction_gui(QDialog,Pipeline):

    def __init__(self):
        
        '''Inputs'''
        self.inputfolder = None
        self.testablefile = None
        self.outputpath = None
        
        '''default user inputs'''
        self.direct_access= True #name of the file is the key in testing table
        self.shift= 1
        self.line = None
        
        '''gui parameter'''
        self.eps_SD = 'Not yet calculated'
        self.key = None
        
        '''time parameter'''
        self.counter = 0
        
        '''Load the ui'''
        QDialog.__init__(self) # Call the inherited classes __init__ method
        self.call = uic.loadUi('../ui/gui_5.ui', self) # Load the .ui file
        
        '''first input: browse'''
        self.call.browse1.clicked.connect(self.open_inputfolder_on_click)#browse folders
        
        '''second input: browse'''
        self.call.browse2.clicked.connect(self.open_tablefile_on_click)#browse folders
        
        '''Third input: browse'''
        self.call.browse3.clicked.connect(self.open_outputfolder_on_click)#browse folders
        
        '''Validate and set in inputs'''
        self.call.validate.clicked.connect(self.validate_inputs_on_click)#check path
        
        '''Check before fitting'''
        self.call.check1.clicked.connect(self.check_before_on_click)#check hysteresis before fitting
        
        '''Check after fitting'''
        self.call.check2.clicked.connect(self.check_after_on_click)#check hysteresis after fitting
        
        '''run script'''
        self.call.run_datareduction.clicked.connect(self.run_on_click)#run the script
        
        '''Parameters settings'''
        self.call.check3.clicked.connect(self.check_inputable)#run the script
        
        '''database'''
        self.call.check4.clicked.connect(self.check_database)#run the script
        
        '''run alll'''
        self.call.run_datareduction_all.clicked.connect(self.run_all_on_click)#run the script
        
        '''database'''
        self.call.add.clicked.connect(self.add_to_database)#run the script
        
        #self.call.show() # Show the GUI
        
    '''run all'''
    def run_all_on_click(self):
        try:
            _import = database_import()
            _import.exec_()
            path = _import.database_path
            if path is not None:
                cnx = sqlite3.connect(path)
                df = pd.read_sql_query("SELECT * FROM specimens", cnx)

                '''data cleaning''' 
                check = ('infolder','testable','strain','stress','extensometer' ,'length', 'diameter','timecycle')
                self.total_rows_before= len(df.index)  
                if (self.total_rows_before==0):   
                     buttonReply = QMessageBox.information(self,'Error message', "Not possible! Empty database")     
                else: 
                    for par in check:
                        self.deleted_index = df[(df[str(par)].map(str)==str('')) | (df[str(par)].map(str)==str('-'))].index
                        df.drop(self.deleted_index,inplace=True)
                        df.reset_index(drop=True,inplace=True)
                    self.total_rows_after= len(df.index)
                    
                    '''data reduction'''
                    self.datareduction_all(df)
            else:
                pass
        except Exception:
            buttonReply = QMessageBox.information(self,'Error message', "Not possible!")
            
    '''add to data base event'''
    def add_to_database(self):
        try:
            specimen = add_directly_todatabase(self)
            specimen.add_todatabase()
        except Exception:
            buttonReply = QMessageBox.information(self,'Inputs error message', "Please validate the paths!")
    
    '''database''' 
    def check_database(self):
        _import = database_import()
        _import.exec_()
        if _import.database_path is not None:
            window = database_window(None,_import.database_path)#None: default header
            window.show()
        
    '''Parameters settings'''
    def check_inputable(self):
        try:
            window_parameters = parameters_window(self)
            window_parameters.exec_()
        except Exception:
            buttonReply = QMessageBox.information(self,'Inputs error message', "Please validate the paths!")
        
    '''position of windows'''
    def pos_window(self,widget,pos='left'):
        display_monitor = 0 # the number of the monitor you want to display your widget

        monitor = QDesktopWidget().screenGeometry(display_monitor)
        #if pos =='right':
        #widget.move( monitor.top(),monitor.right())
        if pos == 'left':
            widget.move(monitor.left(), monitor.top())
        #widget.showFullScreen()
        return widget
        
    '''first input: brows'''
    def open_inputfolder_on_click(self):
        foldername = QFileDialog.getExistingDirectory(self, 'Select directory')
        self.call.lineEdit1.setText(foldername)
        
    '''second input: browse'''
    def open_tablefile_on_click(self):
        filename = QFileDialog.getOpenFileName(self, 'Select file',filter="Excel(*.xls *.xlsx)")
        self.call.lineEdit2.setText(filename[0])
            
    '''Third input: browse'''
    def open_outputfolder_on_click(self):
        foldername = QFileDialog.getExistingDirectory(self, 'Select directory')
        self.call.lineEdit3.setText(foldername)
        
    '''validate'''
    @pyqtSlot()
    def validate_inputs_on_click(self):
        
        path = self.call.lineEdit1.text()
        if not (os.path.isdir(path)):
            buttonReply = QMessageBox.critical(self, 
            'Inputs error message', "The first inserted path should be an existing folder! Please use the browser.")
        else:
            self.inputfolder = self.call.lineEdit1.text()
            
        path = self.call.lineEdit2.text()
        if not (os.path.exists(path)):
            buttonReply = QMessageBox.critical(self, 
            'Inputs error message', "The second path should be an existing 'xls' file! Please use the browser.")
        else:
            self.testablefile = self.call.lineEdit2.text()
            
        path = self.call.lineEdit3.text()
        if not (os.path.isdir(path)):
            buttonReply = QMessageBox.critical(self, 
            'Inputs error message', "The Third path should be an existing folder! Please use the browser.")
        else:
            self.outputpath = self.call.lineEdit3.text() 
            
        if(self.inputfolder is not None and self.testablefile is not None and self.outputpath is not None) :
            self.validate = validate_thread(self)
            self.validate.finished.connect(self.details)
            self.validate.start()

            self.window_progressbar_validate = Progress_bar()
            self.window_progressbar_validate.setWindowTitle('Loading...')
            self.progress = self.window_progressbar_validate.call.progressBar
            self.progress.setRange(0,0)
            self.abort = self.window_progressbar_validate.call.abort
            self.abort.clicked.connect(self.stop_thread_validate)

    def details(self):
        #self.stop_timer()
        self.counter = 0
        self.line = self.validate.specimen.line
        self.window_progressbar_validate.close()
        self.key = self.key
        buttonReply = QMessageBox.information(self, 
        'Inputs error message', "All inputs are valid and data is loaded: \n \n" + 
            '* This is specimen: '+str(self.key) +'\n'+
            '* The specimen folder contains "'+ str(self.n_parts)+'" .txt file. \n'+
            '* The rowdata contains '+str(self.rows) +' lines'+' or '+str(round(self.cycles,2))+' cycles.')
            #print(self.line)
            
    '''check before fitting'''
    def check_before_on_click(self):
        #self.call.hide()
        if(self.inputfolder is not None and self.testablefile is not None):# and self.outputpath is not None):
                    
            '''condition'''
            if (len(glob.glob(os.path.join(self.inputfolder, '*.txt')))>0):
                self.midata_hysteresis(plot =False,save = False)

                if self.test_cycles =='-':
                    half = False
                    percentage = '10%'
                else:
                    half = True
                    percentage = '50%'

                buttonReply = QMessageBox.information(self,' Raw data processing information', 
                '* This is hysteresis of specimen: '+str(self.key) +'\n'+
                '* The specimen folder contains "'+ str(self.n_parts)+'" .txt file. \n'+
                '* '+str(round(self.offset,2))+' cycles were removed (no load cycles)! \n'+
                '* The rowdata contains '+str(self.rows) +' lines'+' or '+str(round(self.cycles,2))+' cycles.\n'+
                '* The hysteresis is corresponding to '+ percentage + ' of the raw data.')

                '''plotting'''
                window2 = Plot_Window(self,False,half,True)
                window2 = self.pos_window(window2)
                window2.show()
            else:
                buttonReply = QMessageBox.critical(self, 
                'Inputs error message', "The folder does not contain any '.txt' file.")
            
        else:
             buttonReply = QMessageBox.critical(self, 
            'Inputs error message', "The first two paths should be valid! Please validate.")
        #self.call.show()
        # self.inputfolder = None
        # self.testablefile = None
        # self.outputpath = None
        
    def plothysteresis(self):
        self.stop_thread_fit()
        
        buttonReply = QMessageBox.information(self,' Fitting information', 
        '* The fitting is done on the '+str(self.start_half)+'th cycle out of '+str(round(self.cycles,2))+' test data cycles. \n'+
        '* The found fitting value [eps_S,eps_D] = '+str(self.eps_SD)+'.\n')
        
        self.fragments = self.sub_ranges()
        self.run_reduced_hyst(False)

        '''plotting'''
        window2 = Plot_Window(self,False,self.half,False)
        window2 = self.pos_window(window2)
        window2.show()

    @pyqtSlot()    
    def check_after_on_click(self):
                
        if(self.inputfolder is not None and self.testablefile is not None):
            '''condition'''
            if (len(glob.glob(os.path.join(self.inputfolder, '*.txt')))>0):
                '''Data processing'''
                self.midata_hysteresis(plot =False,save = False)

                if self.test_cycles =='-':
                    self.half = False
                    percentage = '10%'
                else:
                    self.half = True
                    percentage = '50%'

                buttonReply = QMessageBox.information(self,' Raw data processing information', 
                '* This is hysteresis of specimen: '+str(self.key) +'\n'+
                '* The specimen folder contains "'+ str(self.n_parts)+'" .txt file. \n'+
                '* '+str(round(self.offset,2))+' cycles were removed (no load cycles)! \n'+
                '* The rowdata contains '+str(self.rows) +' lines'+' or '+str(round(self.cycles,2))+' cycles.\n'+
                '* The hysteresis is corresponding to '+ percentage + ' of the crack initiation.')

                '''fitting'''
                buttonReply = QMessageBox.information(self,' Fitting information', 
                ' Click to start fitting! This might take few seconds',
                QMessageBox.Ok|QMessageBox.Cancel, QMessageBox.Ok)

                if(buttonReply==QMessageBox.Ok):
                    
                    self.fit = fitting_thread(self)
                    self.fit.finished.connect(self.plothysteresis)
                    self.fit.start()
                                        
                    self.window_progressbar_fit = Progress_bar()
                    self.window_progressbar_fit.setWindowTitle('Fitting progressing')
                    self.progress = self.window_progressbar_fit.call.progressBar
                    self.progress.setRange(0,0)
                    self.abort = self.window_progressbar_fit.call.abort
                    self.abort.clicked.connect(self.stop_thread_fit)

                else:
                    pass
            else:
                buttonReply = QMessageBox.critical(self, 
                'Inputs error message', "The folder does not contain any '.txt' file.")

        else:
             buttonReply = QMessageBox.critical(self, 
            'Inputs error message', "All paths should be valid! Please validate.")

    @pyqtSlot()    
    def run_on_click(self):
                
        if(self.inputfolder is not None and self.testablefile is not None and self.outputpath is not None):
            '''condition'''
            if (len(glob.glob(os.path.join(self.inputfolder, '*.txt')))>0):
                '''Data processing'''
                self.midata_hysteresis(plot =False,save = False)

                if self.test_cycles =='-':
                    half = False
                    percentage = '10%'
                else:
                    half = True
                    percentage = '50%'

                buttonReply = QMessageBox.information(self,' Raw data processing information', 
                '* This is hysteresis of specimen: '+str(self.key) +'\n'+
                '* The specimen folder contains "'+ str(self.n_parts)+'" .txt file. \n'+
                '* '+str(round(self.offset,2))+' cycles were removed (no load cycles)! \n'+
                '* The rowdata contains '+str(self.rows) +' lines'+' or '+str(round(self.cycles,2))+' cycles.\n'+
                '* The hysteresis is corresponding to '+ percentage + ' of the raw data.')

                '''fitting'''
                buttonReply = QMessageBox.information(self,' Fitting information', 
                ' Click to start fitting! This might take few seconds',
                QMessageBox.Ok|QMessageBox.Cancel, QMessageBox.Ok)

                if(buttonReply==QMessageBox.Ok):
                    
                    self.fit = fitting_thread(self)
                    self.fit.finished.connect(self.datareduction)
                    self.fit.start()
                                        
                    self.window_progressbar_fit = Progress_bar()
                    self.window_progressbar_fit.setWindowTitle('Fitting progressing')
                    self.progress = self.window_progressbar_fit.call.progressBar
                    self.progress.setRange(0,0)
                    self.abort = self.window_progressbar_fit.call.abort
                    self.abort.clicked.connect(self.stop_thread_fit)

                else:
                    pass
            else:
                buttonReply = QMessageBox.critical(self, 
                'Inputs error message', "The folder does not contain any '.txt' file.")

        else:
             buttonReply = QMessageBox.critical(self, 
            'Inputs error message', "All paths should be valid! Please validate.")
    
    @pyqtSlot()
    def datareduction(self):
        
        self.stop_thread_fit()
        
        '''run data reduction'''
        buttonReply = QMessageBox.information(self,' Data reduction information', 
        'Fitting achieved! Now click to start data reduction!',
        QMessageBox.Ok|QMessageBox.Cancel, QMessageBox.Ok)

        if(buttonReply==QMessageBox.Ok): 
            '''one thread for counting seconds'''
            self.timer = QTimer()
            self.timer.setInterval(1000)
            self.timer.timeout.connect(self.recurring_timer)
            self.timer.start()
            ''' the second one for data reduction'''
            self.calc = reduc_thread(self)
            self.calc.result.connect(self.onCountChanged)
            self.calc.finished.connect(self.stop_timer)
            self.calc.start()

            self.window_progressbar = Progress_bar()
            self.window_progressbar.setWindowTitle('Data reduction progressing')
            self.progress = self.window_progressbar.call.progressBar
            self.time_left = self.window_progressbar.call.label
            
            self.abort = self.window_progressbar.call.abort
            self.abort.clicked.connect(self.stop_thread)   
            
        else:
            pass
        
    @pyqtSlot()
    def datareduction_all(self,dataframe = None):
        deleted = self.total_rows_before - self.total_rows_after
        if (deleted>0):
            msg = 'The database is lacking some parameters and '+ str(deleted) +' are not considered! \n Click ok to start data reduction!'
        else:
            msg = 'The datareduction of the chosen database is about to start! \n Click ok to start!'
        '''run data reduction'''
        buttonReply = QMessageBox.information(self,' Data reduction information', 
        msg,QMessageBox.Ok|QMessageBox.Cancel, QMessageBox.Ok)

        if(buttonReply==QMessageBox.Ok): 
            '''one thread for counting seconds'''
            self.timer = QTimer()
            self.timer.setInterval(1000)
            self.timer.timeout.connect(self.recurring_timer)
            self.timer.start()
            ''' the second one for data reduction'''
            self.calc = runall_thread(dataframe)
            self.calc.result.connect(self.onCountChanged_all)
            self.calc.finished.connect(self.stop_timer)
            self.calc.start()

            self.window_progressbar = Progress_bar()
            self.window_progressbar.setWindowTitle('Data reduction progressing')
            self.progress = self.window_progressbar.call.progressBar
            self.time_left = self.window_progressbar.call.label
            
            self.abort = self.window_progressbar.call.abort
            self.abort.clicked.connect(self.stop_thread)   
            
        else:
            pass
         
    '''Threading'''
    def stop_timer(self):
        self.window_progressbar.close()
        self.timer.stop()
        self.counter =0
        
    def stop_thread(self):
        #self.window_progressbar.close()
        self.calc.stop()
        self.counter =0
    
    def onCountChanged(self, value):
        self.progress.setValue(value)
        
        if (value > 1 and value <100):
            current_estimation = round(100*self.counter/value,2) - round(self.counter,2)
            if (current_estimation <60.1):
                self.time_left.setText(str(self.counter)+' seconds has passed! and the estimated left time is ~ "'
                                   + str(round(current_estimation/1.,2))+ '" seconds.')
            elif (current_estimation < 3600):
                self.time_left.setText(str(self.counter)+' seconds has passed! and the estimated left time is ~ "'
                                   + str(round(current_estimation/60.,2))+ '" minutes.') 
            else:
                self.time_left.setText(str(self.counter)+' seconds has passed! and the estimated left time is ~ "'
                                   + str(round(current_estimation/3600.,2))+ '" hours.') 
   
        elif value > 99:
            self.time_left.setText('Data reduction is done and exporting takes few additionnal seconds...')
            self.counter = 0
            
    def onCountChanged_all(self, received_object):

        value = int(received_object[0])
        title = str(received_object[1])
        index = int(received_object[2]+1)
        total = int(received_object[3])
        self.progress.setValue(value)
        self.window_progressbar.setWindowTitle('Data reduction process of: '+title)
        self.infos = self.window_progressbar.call.label1
        
        self.infos.setText(str(index-1)+ ' are done!'+str(total-index+1)+' specimens are still to run (this is the'+ str(index)+'th out of '+str(total)+')')
        
        if (value > 1 and value <100): 
            current_estimation = round(100*self.counter/value,2) - round(self.counter,2)
            if (current_estimation <60.1):
                self.time_left.setText(str(self.counter)+' seconds has passed! and the estimated left time is ~ "'
                                   + str(round(current_estimation/1.,2))+ '" seconds.')
            elif (current_estimation < 3600):
                self.time_left.setText(str(self.counter)+' seconds has passed! and the estimated left time is ~ "'
                                   + str(round(current_estimation/60.,2))+ '" minutes.') 
            else:
                self.time_left.setText(str(self.counter)+' seconds has passed! and the estimated left time is ~ "'
                                   + str(round(current_estimation/3600.,2))+ '" hours.') 
   
        elif value > 99:
            self.time_left.setText('Data reduction is done and exporting takes few additionnal seconds...')
            self.counter = 0
                    
    def stop_thread_fit(self):
        self.fit.stop()
        self.window_progressbar_fit.close()
        
    def stop_thread_validate(self):
        self.counter = 0
        self.validate.stop()
        self.window_progressbar_validate.close()
    
    def recurring_timer(self):
        self.counter +=1
                
#test            
def main():
    app = QApplication(sys.argv) # Create an instance of QApplication
    app.setStyle(QStyleFactory.create('Fusion')) # otherwise colors change won't work on windows style.
    window = datareduction_gui() 
    # Create an instance of our class
    #app.aboutToQuit.connect(app.deleteLater)
    sys.exit(app.exec_()) # Start the application
#main()

# # Funny msg

class funny_msg(QDialog):

    def __init__(self):
        
        '''Load the ui'''
        QDialog.__init__(self) # Call the inherited classes __init__ method
        self.call = uic.loadUi('../ui/../ui/Funny_msg.ui', self) # Load the .ui file
        #self.show()
        
## Devloper information

class developer_infos(QDialog):

    def __init__(self):
        
        '''Load the ui'''
        QDialog.__init__(self) # Call the inherited classes __init__ method
        self.call = uic.loadUi('../ui/coordinates.ui', self) # Load the .ui file
        #self.show()

## Main window
class main_window_gui(QMainWindow):
    def __init__(self):
        self.window = None
        '''Load the ui'''
        QMainWindow.__init__(self) # Call the inherited classes __init__ method
        self.call = uic.loadUi('../ui/tool.ui', self) # Load the .ui file
        
        '''Douglas-Peucker-Algo : menubar-->menudata'''
        self.algo1 = self.call.action_algo
        self.algo1.triggered.connect(self.Douglas_Peucker_Algo)
        
        '''other method : menubar-->menudata'''
        self.algo2 = self.call.action_more
        self.algo2.triggered.connect(self.Algo2)
        
        '''ppt : menubar-->menuhelp'''
        self.ppt = self.call.action_ppt
        self.ppt.triggered.connect(self.GUI_ppt)
        
        '''code doc : menubar-->menuhelp'''
        self.doc = self.call.action_doc
        self.doc.triggered.connect(self.code_doc)
        
        '''developer : menubar-->menuhelp'''
        self.contact = self.call.action_contact
        self.contact.triggered.connect(self.coordinates)

        self.show()
        
    def Douglas_Peucker_Algo(self):
        #killing old processes
        if self.window is not None:
            #self.window.close()
            pass
        else:
            pass
        #or using setenables(false) after one use
        self.window = datareduction_gui()  
        self.setCentralWidget(self.window)
        self.call.action_algo.setEnabled(False)
        #window.show()

    def Algo2(self):
        window = funny_msg() 
        window.show()

    def GUI_ppt(self):
        try:
            abspath = os.path.abspath('../docs')
            name = 'Datareduction guideline.pptx'
            os.startfile(os.path.join(abspath, name))
            
        except Exception:
             buttonReply = QMessageBox.critical(self, 
            'Error message', "File could not be open")
            
    def code_doc(self):
        try:
            abspath = os.path.realpath('../docs/_build/html/index.html')
            webbrowser.open(abspath)
            
        except Exception:
             buttonReply = QMessageBox.critical(self, 
            'Error message', "File could not be open")
                
    def coordinates(self):
       
        window = developer_infos() 
        window.show()
        #self.setCentralWidget(self.window)

#test
def main():
    app = QApplication(sys.argv) # Create an instance of QApplication
    app.setStyle(QStyleFactory.create('Fusion')) # otherwise colors change won't work on windows style.
    window = main_window_gui() 
    # Create an instance of our class
    #app.aboutToQuit.connect(app.deleteLater)
    app.exec_() # Start the application

main()

